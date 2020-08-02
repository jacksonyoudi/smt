# -*- coding: utf-8 -*-

from openpyxl import load_workbook

if __name__ == '__main__':
    wb = load_workbook(filename="/Users/youdi/Desktop/20200721.xlsx")
    sheet = wb["产能总表"]
    for i in sheet.iter_rows():
        if i[2].value and i[4].value and i[5].value:
            print(i[2].value, i[4].value, i[5].value, *[i[j].value for j in range(44, 67)])

# import pandas as pd

# frame = pd.read_excel("/Users/youdi/Documents/20200721.xlsm")
